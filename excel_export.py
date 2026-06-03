from datetime import date
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import database

_GROEN = "FF2E7D32"
_LICHTGROEN = "FFE8F5E9"
_GRIJS = "FFF5F5F5"
_WIT = "FFFFFFFF"


def _rand(dik=False):
    stijl = "medium" if dik else "thin"
    z = Side(style=stijl)
    return Border(left=z, right=z, top=z, bottom=z)


def _koptekst(cel, tekst, vet=True, achtergrond=_GROEN, tekstkleur="FFFFFFFF"):
    cel.value = tekst
    cel.font = Font(bold=vet, color=tekstkleur, size=11)
    cel.fill = PatternFill("solid", fgColor=achtergrond)
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cel.border = _rand(dik=True)


def _cel(cel, waarde, achtergrond=_WIT, vet=False, getal=False):
    cel.value = waarde
    cel.font = Font(bold=vet, size=10)
    cel.fill = PatternFill("solid", fgColor=achtergrond)
    cel.alignment = Alignment(horizontal="right" if getal else "left", vertical="center")
    cel.border = _rand()


def _pas_kolombreedte_aan(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


def _tabblad_detail(wb: openpyxl.Workbook, registraties: list[dict], van: date, tot: date):
    ws = wb.active
    ws.title = "Alle registraties"

    ws.row_dimensions[1].height = 22
    koppen = ["Datum", "Tijdstip", "Telefoonnummer", "Productnaam", "Variant", "Kilo's"]
    for col, kop in enumerate(koppen, start=1):
        _koptekst(ws.cell(1, col), kop)

    for rij_nr, reg in enumerate(registraties, start=2):
        bg = _LICHTGROEN if rij_nr % 2 == 0 else _WIT
        _cel(ws.cell(rij_nr, 1), reg["datum"], bg)
        _cel(ws.cell(rij_nr, 2), reg["tijdstip"], bg)
        _cel(ws.cell(rij_nr, 3), reg["telefoon"], bg)
        _cel(ws.cell(rij_nr, 4), reg["productnaam"], bg)
        _cel(ws.cell(rij_nr, 5), reg["variant"], bg)
        _cel(ws.cell(rij_nr, 6), reg["kilo"], bg, getal=True)

    if registraties:
        totaal_rij = len(registraties) + 2
        _cel(ws.cell(totaal_rij, 5), "TOTAAL", _GRIJS, vet=True)
        totaal = sum(r["kilo"] for r in registraties)
        _cel(ws.cell(totaal_rij, 6), round(totaal, 2), _GRIJS, vet=True, getal=True)

    _pas_kolombreedte_aan(ws)
    ws.freeze_panes = "A2"


def _tabblad_samenvatting(wb: openpyxl.Workbook, registraties: list[dict], van: date, tot: date):
    ws = wb.create_sheet("Wekelijks totaal")

    ws.cell(1, 1).value = f"Periode: {van.strftime('%d-%m-%Y')} t/m {tot.strftime('%d-%m-%Y')}"
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 20

    ws.row_dimensions[2].height = 22
    koppen = ["Productnaam", "Variant", "Totaal kilo's", "Aantal registraties"]
    for col, kop in enumerate(koppen, start=1):
        _koptekst(ws.cell(2, col), kop)

    totalen: dict[tuple, list] = defaultdict(list)
    for reg in registraties:
        totalen[(reg["productnaam"], reg["variant"])].append(reg["kilo"])

    gesorteerd = sorted(totalen.items(), key=lambda x: (-sum(x[1]), x[0][0]))
    for rij_nr, ((naam, variant), kilos) in enumerate(gesorteerd, start=3):
        bg = _LICHTGROEN if rij_nr % 2 == 0 else _WIT
        _cel(ws.cell(rij_nr, 1), naam, bg)
        _cel(ws.cell(rij_nr, 2), variant, bg)
        _cel(ws.cell(rij_nr, 3), round(sum(kilos), 2), bg, getal=True)
        _cel(ws.cell(rij_nr, 4), len(kilos), bg, getal=True)

    if totalen:
        totaal_rij = len(totalen) + 3
        _cel(ws.cell(totaal_rij, 2), "TOTAAL", _GRIJS, vet=True)
        _cel(ws.cell(totaal_rij, 3), round(sum(sum(v) for v in totalen.values()), 2), _GRIJS, vet=True, getal=True)
        _cel(ws.cell(totaal_rij, 4), sum(len(v) for v in totalen.values()), _GRIJS, vet=True, getal=True)

    _pas_kolombreedte_aan(ws)
    ws.freeze_panes = "A3"


def genereer_rapport(van: date, tot: date) -> str:
    """Genereert een .xlsx-bestand en retourneert het pad."""
    registraties = database.haal_registraties_op(van, tot)

    wb = openpyxl.Workbook()
    _tabblad_detail(wb, registraties, van, tot)
    _tabblad_samenvatting(wb, registraties, van, tot)

    week_nr = van.isocalendar().week
    jaar = van.year
    pad = f"verspilling_week_{week_nr:02d}_{jaar}.xlsx"
    wb.save(pad)
    return pad
